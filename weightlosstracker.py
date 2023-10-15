import openpyxl
from datetime import datetime
import os

def create_workbook(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["Date & Time", "Body Weight (lbs)", "Body Fat %", "Age", "Height (ft'inch\")", "BMR"]
    ws.append(headers)
    wb.save(filename)

def lbs_to_kg(weight_in_lbs):
    return weight_in_lbs * 0.453592

def ftinch_to_cm(height_str):
    # Converting imperial height to cm, for the BMR equation
    feet, inches = map(int, height_str.replace("\"", "").split("'"))
    total_inches = (feet * 12) + inches
    return total_inches * 2.54

def add_entry(filename, weight, body_fat=None, age=None, height_str=None):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # If the row doesn't have age and height, use the last available entry for calculation
    if not age or not height_str:
        last_row = ws.max_row
        if not age:
            age = ws.cell(row=last_row, column=4).value
        if not height_str:
            height_str = ws.cell(row=last_row, column=5).value

    weight_kg = lbs_to_kg(weight)
    height_cm = ftinch_to_cm(height_str)

    # Calculate BMR for male (modify for female if needed)
    bmr = 10 * weight_kg + 6.25 * height_cm - 5 * age + 5

    data = [datetime.now(), weight, body_fat, age, height_str, bmr]
    ws.append(data)

    wb.save(filename)

def set_goal(filename, target_weight):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    ws['H1'] = "Target Weight (lbs)"
    ws['H2'] = target_weight

    wb.save(filename)

def main():

    filename = "weight_loss_tracker.xlsx"
    
    # Check if the file already exists. If not, create a new workbook.
    # Won't upload my own file to Github repo, lol -- Don't want to publish my bodyweight for the world to see
    if not os.path.exists(filename):
        create_workbook(filename)


    while True:
        action = input("Choose action (add / goal / exit): ").lower()

        if action == "add":
            weight = float(input("Enter Body Weight (lbs): "))
            body_fat = input("Enter Body Fat % (leave blank if not available): ")
            age = input("Enter Age (leave blank if not available): ")
            height_str = input("Enter Height (ft'inch\") (e.g. 5'10\") (leave blank if not available): ")

            body_fat = float(body_fat) if body_fat else None
            age = int(age) if age else None

            add_entry(filename, weight, body_fat, age, height_str)
        elif action == "goal":
            target_weight = float(input("Enter Target Weight (lbs): "))
            set_goal(filename, target_weight)
        elif action == "exit":
            break

if __name__ == "__main__":
    main()
